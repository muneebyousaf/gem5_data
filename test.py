import re 
import os.path
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font


stream_path="/home/muneeb/gem5/processing/load/stream/"
spec_path="/home/muneeb/gem5/processing/load/spec/"

directory_contents = os.listdir(stream_path)
for dir1 in directory_contents:
	load_path= stream_path+dir1
	print(load_path);
 
	

file_exists = os.path.exists('attack_traces.xlsx')
if file_exists == False:

	print (" file does not exist creating a new one \n");
	workbook = Workbook()
	workbook.create_sheet("traces", 0) # insert at first position
	workbook.save(filename='attack_traces.xlsx')
	workbook = load_workbook(filename="attack_traces.xlsx")

else: 
	print(" file exist \n");
	workbook = load_workbook(filename="attack_traces.xlsx")

workbook.active=0
sheet=workbook.active;

states = open("./m5out/stats.txt", "r")

Lines = states.readlines()

col_count=0 ; 

line_list=[]
offset=1
load_description=  "Description about the load"
sheet.cell(row=offset, column=1).value = load_description
offset += 1;
label= " gem5 state parameter Name"
comment=" Parameter Description"
data1 = "value 1"
data2 = "value 2"
data3 = "value 3"
sheet.cell(row=offset, column=1).value = label
sheet.cell(row=offset, column=2).value = data1
sheet.cell(row=offset, column=3).value = data2
sheet.cell(row=offset, column=4).value = data3
sheet.cell(row=offset, column=5).value = comment
offset += 1;
row_count=offset;
for line in Lines:
	line=line.strip()
	line_list = re.split(r'\s{2,}|#',line)
	a = len(line_list);
	if a  > 1:
		row_count +=1;
	#	print(a)
		for sublist in line_list:
			col_count +=1;
			if col_count <  a: 
				sheet.cell(row=row_count, column=col_count).value = sublist
			else:
				col_count=5
				sheet.cell(row=row_count, column=col_count).value = sublist

		col_count =0;


workbook.save(filename='attack_traces.xlsx')
	

	#print("Line{}: {}".format(count, line))
#sheet = workbook.active
#sheet["A1"] = "trace label"
#sheet["B1"] = "FF"

#sheet['B1'].font = Font(bold=True)


#for x in range(1,101):
#...        for y in range(1,101):
#...            ws.cell(row=x, column=y)
#workbook.save(filename='attack_traces.xlsx')


