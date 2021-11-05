import sys 
import re
import subprocess
import os.path
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font



file_exists = os.path.exists('attack_traces.xlsx')
if file_exists == False:

        print (" file does not exist creating a new one \n");
        workbook = Workbook()
        workbook.create_sheet("traces", 0) # insert at first position
        workbook.save(filename='attack_traces.xlsx')
        workbook = load_workbook(filename="attack_traces.xlsx")

else:
        print(" file exist \n");
        workbook = load_workbook(filename="attack_traces.xlsx")

workbook.active=0
sheet = workbook.active
#eet["A1"] = "trace label"
#heet["B1"] = "FF"
#heet['B1'].font = Font(bold=True)


#or x in range(1,101):
#..        for y in range(1,101):
#..            ws.cell(row=x, column=y)
#orkbook.save(filename='attack_traces.xlsx')

#rint(workbook.sheetnames)


stream_path="/home/muneeb/gem5/processing/load/stream/"
spec_path="/home/muneeb/gem5/processing/load/spec/"

path = spec_path # name of folder to execute 

directory_contents = os.listdir(path)
#directory_contents = os.listdir(stream_path)
app_count = 0 ; 
max_col = 5 ; 
for dir1 in directory_contents:
	load_path= path+dir1
        
	gem5= "/home/muneeb/gem5/build/X86/gem5.opt "
	config= sys.argv[1]+" " 
	subprocess.call( gem5 + config+ load_path, shell=True)

	app_offset = app_count * max_col;

	app_count+=1

	

	states = open("./m5out/stats.txt", "r")

	Lines = states.readlines()

	col_count=0 ; 

	line_list=[]
	offset=1
	load_description= load_path
	sheet.cell(row=offset, column=1 + app_offset).value = load_description
	offset += 1;
	label= " gem5 state parameter Name"
	comment=" Parameter Description"
	data1 = "value 1"
	data2 = "value 2"
	data3 = "value 3"
	sheet.cell(row=offset, column=1+ app_offset).value = label
	sheet.cell(row=offset, column=2 +app_offset).value = data1
	sheet.cell(row=offset, column=3 + app_offset).value = data2
	sheet.cell(row=offset, column=4 + app_offset).value = data3
	sheet.cell(row=offset, column=5 + app_offset).value = comment
	row_count=offset;
	for line in Lines:
		line=line.strip()
		line_list = re.split(r'\s{2,}|#',line)
		list_element_count = len(line_list);
		if list_element_count  > 1:
			row_count +=1;
			for sublist in line_list:
				col_count +=1;
				if col_count <  list_element_count: 
					sheet.cell(row=row_count, column=col_count+app_offset).value = sublist
				else:
					col_count=5
					sheet.cell(row=row_count, column=col_count + app_offset).value = sublist

			col_count =0;


workbook.save(filename='attack_traces.xlsx')
	

